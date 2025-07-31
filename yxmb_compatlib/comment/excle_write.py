# -*- coding: utf8 -*-

import logging
import csv
import os
from openpyxl import load_workbook, Workbook


def excel_append2(file_path, column_headers, contents):
    # 检查输入数据长度一致性
    if len(column_headers) != len(contents):
        raise ValueError(
            'The number of column headers must match the number of contents.'
        )

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        try:
            # 尝试加载现有的Excel文件
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            # 文件不存在，创建新文件并写入表头
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(column_headers)  # 写入表头

        # 将新内容写入新的一行
        sheet.append(contents)
        workbook.save(file_path)
        logging.info(f'已将数据追加至 {file_path}')

    elif file_extension == '.csv':
        # 检查CSV文件是否已存在
        file_exists = os.path.exists(file_path)

        with open(file_path, mode='a', encoding='utf-8', newline='') as file:
            writer = csv.writer(file)

            if not file_exists:
                # 如果文件不存在，写入表头
                writer.writerow(column_headers)

            # 写入内容
            writer.writerow(contents)
        logging.info(f'已将数据追加至 {file_path}')

    else:
        raise ValueError('Unsupported file format. Only .xlsx and .csv are supported.')
