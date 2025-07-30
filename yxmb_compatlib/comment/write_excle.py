import csv
import logging
import os
from openpyxl import load_workbook, Workbook


def excel_append(file_path, column_header1, content1, column_header2, content2):
    """
    功能：向指定的Excel（.xlsx）或CSV文件的指定两列表头下追加内容。
    - 如果对应列有空白单元格，则填入内容；
    - 如果没有空白，则在文件末尾新增一行并写入内容。
    参数:
        file_path: 文件路径（支持.xlsx和.csv）
        column_header1: 第一个目标列的表头
        content1: 第一个目标列要写入的内容
        column_header2: 第二个目标列的表头
        content2: 第二个目标列要写入的内容
    """
    # 检查文件扩展名
    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]  # 获取第一行的所有表头

        # 确定要操作的列索引
        column_index1 = headers.index(column_header1) + 1
        column_index2 = headers.index(column_header2) + 1

        # 从第一行开始，向下查找直到找到两个空白单元格中的一个
        for row in range(1, sheet.max_row + 1):
            if (
                not sheet.cell(row=row, column=column_index1).value
                or not sheet.cell(row=row, column=column_index2).value
            ):
                # 找到空白单元格，跳出循环
                break
        else:
            # 如果所有单元格都不为空，则在最后一行之后添加新行
            row = sheet.max_row + 1

        # 如果内容是列表，则将其转换为字符串
        if isinstance(content1, list):
            content1 = '\n'.join(content1)
        if isinstance(content2, list):
            content2 = '\n'.join(content2)

        # 写入内容到找到的空白单元格
        sheet.cell(row=row, column=column_index1).value = content1
        sheet.cell(row=row, column=column_index2).value = content2

        workbook.save(file_path)

    elif file_extension == '.csv':
        # 读取现有数据
        with open(file_path, mode='r', encoding='utf-8', newline='') as file:
            reader = list(csv.reader(file))

        headers = reader[0]  # 获取CSV文件的第一行作为表头

        # 确定要操作的列索引
        if column_header1 not in headers:
            raise ValueError(f"Column header '{column_header1}' not found in CSV file.")
        if column_header2 not in headers:
            raise ValueError(f"Column header '{column_header2}' not found in CSV file.")

        column_index1 = headers.index(column_header1)
        column_index2 = headers.index(column_header2)

        # 查找第一个空白单元格
        row_to_write = None
        for i, row in enumerate(reader):
            if len(row) <= max(column_index1, column_index2):
                row.extend([''] * (max(column_index1, column_index2) - len(row) + 1))

            if not row[column_index1] or not row[column_index2]:
                row_to_write = i
                break

        if row_to_write is not None:
            # 找到空白行，写入内容
            reader[row_to_write][column_index1] = content1
            reader[row_to_write][column_index2] = content2
        else:
            # 没有空白行，新增一行
            new_row = [''] * (len(headers))
            new_row[column_index1] = content1
            new_row[column_index2] = content2
            reader.append(new_row)

        # 写回csv文件
        with open(file_path, mode='w', encoding='utf-8', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(reader)

    else:
        raise ValueError('Unsupported file format. Only .xlsx and .csv are supported.')


# 调用函数示例
# excel_append("执行结果/查询结果.csv", "身份证号", sfzh, "签约信息", text)


def excel_append2(file_path, column_headers, contents):
    """
    功能：向指定的Excel（.xlsx）或CSV文件追加一整行数据。
    - 如果文件不存在则新建文件并写入表头；
    - 如果文件已存在则直接追加数据行。
    参数:
        file_path: 文件路径（支持.xlsx和.csv）
        column_headers: 列表，表头
        contents: 列表，要写入的数据内容
    """
    # 检查输入数据长度一致性
    if len(column_headers) != len(contents):
        raise ValueError(
            'The number of column headers must match the number of contents.'
        )

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        try:
            # 尝试加载现有的Excel文件
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            # 文件不存在，创建新文件并写入表头
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(column_headers)  # 写入表头

        # 将新内容写入新的一行
        sheet.append(contents)
        workbook.save(file_path)
        logging.info(f'已将数据追加至 {file_path}')

    elif file_extension == '.csv':
        # 检查CSV文件是否已存在
        file_exists = os.path.exists(file_path)

        with open(file_path, mode='a', encoding='utf-8', newline='') as file:
            writer = csv.writer(file)

            if not file_exists:
                # 如果文件不存在，写入表头
                writer.writerow(column_headers)

            # 写入内容
            writer.writerow(contents)
        logging.info(f'已将数据追加至 {file_path}')

    else:
        raise ValueError('Unsupported file format. Only .xlsx and .csv are supported.')
