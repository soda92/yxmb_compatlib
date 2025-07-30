import logging
from openpyxl import load_workbook


def check_and_delete_last_row(file_path):
    # 加载工作簿
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取最后一行的行号
    last_row = sheet.max_row

    # 检查第一列的最后一行是否为空，同时第三列的最后一行是否有数据
    if (
        sheet.cell(row=last_row, column=1).value is None
        and sheet.cell(row=last_row, column=3).value is not None
    ):
        # 条件满足，删除整行
        sheet.delete_rows(last_row)
        # 保存工作簿
        workbook.save(file_path)
        logging.info(f' {last_row} 被清空')

    else:
        # 条件不满足，不执行操作
        logging.info(f'{last_row} 行无需清空')
